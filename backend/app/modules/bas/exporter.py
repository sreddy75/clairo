"""BAS Working Paper Exporter.

Generates PDF and Excel exports of BAS working papers for accountant review.
"""

import io
from datetime import UTC, datetime
from decimal import Decimal
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

if TYPE_CHECKING:
    from app.modules.bas.models import BASCalculation, BASPeriod, BASSession


def format_currency(value: Decimal | str | None) -> str:
    """Format a value as AUD currency."""
    if value is None:
        return "$0.00"
    if isinstance(value, str):
        value = Decimal(value)
    return f"${value:,.2f}"


def format_date(dt: datetime | None) -> str:
    """Format a datetime for display."""
    if dt is None:
        return "-"
    return dt.strftime("%d %b %Y")


class BASWorkingPaperExporter:
    """Exports BAS working papers to PDF and Excel formats."""

    def __init__(
        self,
        session: "BASSession",
        period: "BASPeriod",
        calculation: "BASCalculation | None",
        organization_name: str,
    ):
        self.session = session
        self.period = period
        self.calculation = calculation
        self.organization_name = organization_name

    def generate_pdf(self) -> bytes:
        """Generate a PDF working paper.

        Returns:
            PDF file as bytes
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=12,
            alignment=1,  # Center
        )
        story.append(Paragraph("Business Activity Statement", title_style))
        story.append(Paragraph("Working Paper", title_style))
        story.append(Spacer(1, 10 * mm))

        # Organization Info
        info_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontSize=11,
            spaceAfter=4,
        )
        story.append(Paragraph(f"<b>Organization:</b> {self.organization_name}", info_style))
        story.append(Paragraph(f"<b>Period:</b> {self.period.display_name}", info_style))
        story.append(
            Paragraph(
                f"<b>Period Dates:</b> {format_date(self.period.start_date)} - {format_date(self.period.end_date)}",
                info_style,
            )
        )
        story.append(Paragraph(f"<b>Due Date:</b> {format_date(self.period.due_date)}", info_style))
        story.append(
            Paragraph(f"<b>Status:</b> {self.session.status.replace('_', ' ').title()}", info_style)
        )
        story.append(Spacer(1, 10 * mm))

        if self.calculation:
            # GST Section
            story.append(Paragraph("<b>GST Calculation</b>", styles["Heading2"]))
            story.append(Spacer(1, 5 * mm))

            gst_data = [
                ["Field", "Description", "Amount"],
                [
                    "G1",
                    "Total Sales (including GST)",
                    format_currency(self.calculation.g1_total_sales),
                ],
                ["G2", "Export Sales", format_currency(self.calculation.g2_export_sales)],
                ["G3", "Other GST-Free Sales", format_currency(self.calculation.g3_gst_free_sales)],
                [
                    "G10",
                    "Capital Purchases",
                    format_currency(self.calculation.g10_capital_purchases),
                ],
                [
                    "G11",
                    "Non-Capital Purchases",
                    format_currency(self.calculation.g11_non_capital_purchases),
                ],
                ["", "", ""],
                ["1A", "GST on Sales", format_currency(self.calculation.field_1a_gst_on_sales)],
                [
                    "1B",
                    "GST on Purchases",
                    format_currency(self.calculation.field_1b_gst_on_purchases),
                ],
                ["", "Net GST Payable/(Refundable)", format_currency(self.calculation.gst_payable)],
            ]

            gst_table = Table(gst_data, colWidths=[30 * mm, 80 * mm, 40 * mm])
            gst_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dbeafe")),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ]
                )
            )
            story.append(gst_table)
            story.append(Spacer(1, 10 * mm))

            # PAYG Section (if applicable)
            if (
                self.calculation.w1_total_wages
                and Decimal(str(self.calculation.w1_total_wages)) > 0
            ):
                story.append(Paragraph("<b>PAYG Withholding</b>", styles["Heading2"]))
                story.append(Spacer(1, 5 * mm))

                payg_data = [
                    ["Field", "Description", "Amount"],
                    [
                        "W1",
                        "Total Salary, Wages and Other Payments",
                        format_currency(self.calculation.w1_total_wages),
                    ],
                    [
                        "W2",
                        "Amount Withheld from Payments",
                        format_currency(self.calculation.w2_amount_withheld),
                    ],
                ]

                payg_table = Table(payg_data, colWidths=[30 * mm, 80 * mm, 40 * mm])
                payg_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7c3aed")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ("TOPPADDING", (0, 0), (-1, -1), 6),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ]
                    )
                )
                story.append(payg_table)
                story.append(Spacer(1, 10 * mm))

            # Summary Section
            story.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
            story.append(Spacer(1, 5 * mm))

            total_payable = Decimal(str(self.calculation.total_payable))
            is_refund = total_payable < 0

            summary_data = [
                ["", "Amount"],
                ["Net GST", format_currency(self.calculation.gst_payable)],
                ["PAYG Withheld", format_currency(self.calculation.w2_amount_withheld)],
                [
                    "Total Payable to ATO" if not is_refund else "Total Refund from ATO",
                    format_currency(abs(total_payable)),
                ],
            ]

            summary_table = Table(summary_data, colWidths=[110 * mm, 40 * mm])
            summary_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 11),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        (
                            "BACKGROUND",
                            (0, -1),
                            (-1, -1),
                            colors.HexColor("#fef3c7")
                            if not is_refund
                            else colors.HexColor("#dbeafe"),
                        ),
                    ]
                )
            )
            story.append(summary_table)
            story.append(Spacer(1, 10 * mm))

            # Metadata
            meta_style = ParagraphStyle(
                "Meta",
                parent=styles["Normal"],
                fontSize=9,
                textColor=colors.grey,
            )
            story.append(
                Paragraph(
                    f"Calculated: {self.calculation.calculated_at.strftime('%d %b %Y %H:%M')}",
                    meta_style,
                )
            )
            story.append(
                Paragraph(
                    f"Data: {self.calculation.invoice_count} invoices, {self.calculation.transaction_count} transactions",
                    meta_style,
                )
            )
        else:
            story.append(Paragraph("No calculation data available.", styles["Normal"]))

        # Footer
        story.append(Spacer(1, 20 * mm))
        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
            alignment=1,
        )
        story.append(
            Paragraph(
                f"Generated by Clairo on {datetime.now(UTC).strftime('%d %b %Y %H:%M')}",
                footer_style,
            )
        )

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel(self) -> bytes:
        """Generate an Excel working paper.

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "BAS Working Paper"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        currency_font = Font(name="Calibri", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 1

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "Business Activity Statement - Working Paper"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        row += 2

        # Organization Info
        info_data = [
            ("Organization:", self.organization_name),
            ("Period:", self.period.display_name),
            (
                "Period Dates:",
                f"{format_date(self.period.start_date)} - {format_date(self.period.end_date)}",
            ),
            ("Due Date:", format_date(self.period.due_date)),
            ("Status:", self.session.status.replace("_", " ").title()),
        ]

        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        if self.calculation:
            # GST Section Header
            ws.merge_cells(f"A{row}:D{row}")
            ws.cell(row=row, column=1, value="GST Calculation").font = Font(bold=True, size=14)
            row += 1

            # GST Table Header
            gst_headers = ["Field", "Description", "Amount"]
            for col, header in enumerate(gst_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            # GST Data
            gst_data = [
                ("G1", "Total Sales (including GST)", self.calculation.g1_total_sales),
                ("G2", "Export Sales", self.calculation.g2_export_sales),
                ("G3", "Other GST-Free Sales", self.calculation.g3_gst_free_sales),
                ("G10", "Capital Purchases", self.calculation.g10_capital_purchases),
                ("G11", "Non-Capital Purchases", self.calculation.g11_non_capital_purchases),
                ("", "", ""),
                ("1A", "GST on Sales", self.calculation.field_1a_gst_on_sales),
                ("1B", "GST on Purchases", self.calculation.field_1b_gst_on_purchases),
                ("", "Net GST Payable/(Refundable)", self.calculation.gst_payable),
            ]

            for field, desc, amount in gst_data:
                ws.cell(row=row, column=1, value=field).border = thin_border
                ws.cell(row=row, column=2, value=desc).border = thin_border
                cell = ws.cell(row=row, column=3, value=float(amount) if amount else 0)
                cell.number_format = '"$"#,##0.00'
                cell.border = thin_border
                cell.font = currency_font
                row += 1

            row += 1

            # PAYG Section (if applicable)
            if (
                self.calculation.w1_total_wages
                and Decimal(str(self.calculation.w1_total_wages)) > 0
            ):
                ws.merge_cells(f"A{row}:D{row}")
                ws.cell(row=row, column=1, value="PAYG Withholding").font = Font(bold=True, size=14)
                row += 1

                # PAYG Header
                payg_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
                for col, header in enumerate(["Field", "Description", "Amount"], 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = payg_fill
                    cell.border = thin_border
                row += 1

                # PAYG Data
                payg_data = [
                    (
                        "W1",
                        "Total Salary, Wages and Other Payments",
                        self.calculation.w1_total_wages,
                    ),
                    ("W2", "Amount Withheld from Payments", self.calculation.w2_amount_withheld),
                ]

                for field, desc, amount in payg_data:
                    ws.cell(row=row, column=1, value=field).border = thin_border
                    ws.cell(row=row, column=2, value=desc).border = thin_border
                    cell = ws.cell(row=row, column=3, value=float(amount) if amount else 0)
                    cell.number_format = '"$"#,##0.00'
                    cell.border = thin_border
                    row += 1

                row += 1

            # Summary Section
            ws.merge_cells(f"A{row}:D{row}")
            ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=14)
            row += 1

            total_payable = Decimal(str(self.calculation.total_payable))
            is_refund = total_payable < 0

            summary_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
            for col, header in enumerate(["", "Amount"], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = summary_fill
                cell.border = thin_border
            row += 1

            summary_data = [
                ("Net GST", self.calculation.gst_payable),
                ("PAYG Withheld", self.calculation.w2_amount_withheld),
                (
                    "Total Payable to ATO" if not is_refund else "Total Refund from ATO",
                    abs(total_payable),
                ),
            ]

            for desc, amount in summary_data:
                ws.cell(row=row, column=1, value=desc).border = thin_border
                cell = ws.cell(row=row, column=2, value=float(amount) if amount else 0)
                cell.number_format = '"$"#,##0.00'
                cell.border = thin_border
                row += 1

            row += 2

            # Metadata
            ws.cell(row=row, column=1, value="Calculated:").font = Font(italic=True, color="808080")
            ws.cell(
                row=row, column=2, value=self.calculation.calculated_at.strftime("%d %b %Y %H:%M")
            ).font = Font(color="808080")
            row += 1
            ws.cell(row=row, column=1, value="Data:").font = Font(italic=True, color="808080")
            ws.cell(
                row=row,
                column=2,
                value=f"{self.calculation.invoice_count} invoices, {self.calculation.transaction_count} transactions",
            ).font = Font(color="808080")

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 20

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

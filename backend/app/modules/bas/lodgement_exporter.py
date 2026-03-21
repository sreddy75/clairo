"""Enhanced BAS Exporter with ATO-compliant lodgement summaries.

Spec 011: Interim Lodgement
"""

import io
from datetime import UTC, datetime
from decimal import ROUND_HALF_UP, Decimal
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.modules.bas.exporter import BASWorkingPaperExporter, format_date
from app.modules.bas.schemas import LodgementField

if TYPE_CHECKING:
    from app.modules.bas.models import BASCalculation, BASPeriod, BASSession


def format_whole_dollars(value: Decimal | str | None) -> str:
    """Format a value as whole dollars (no cents) for ATO compliance."""
    if value is None:
        return "$0"
    if isinstance(value, str):
        value = Decimal(value)
    rounded = int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return f"${rounded:,}"


class LodgementExporter(BASWorkingPaperExporter):
    """Enhanced BAS exporter with ATO-compliant lodgement summaries.

    Extends the base BASWorkingPaperExporter to add:
    - Lodgement summary section to PDF exports
    - Lodgement Summary sheet to Excel exports
    - All amounts rounded to whole dollars per ATO requirements
    """

    def __init__(
        self,
        session: "BASSession",
        period: "BASPeriod",
        calculation: "BASCalculation | None",
        organization_name: str,
        abn: str | None = None,
        approved_by_name: str | None = None,
        lodged_by_name: str | None = None,
    ):
        super().__init__(session, period, calculation, organization_name)
        self.abn = abn
        self.approved_by_name = approved_by_name
        self.lodged_by_name = lodged_by_name

    def _round_to_whole_dollars(self, amount: Decimal | None) -> int:
        """Round amount to whole dollars per ATO requirements."""
        if amount is None:
            return 0
        return int(Decimal(str(amount)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

    def _build_lodgement_summary_data(self) -> list[LodgementField]:
        """Build list of BAS fields formatted for ATO portal entry.

        Returns fields in ATO form order: G1, G2, G3, G10, G11, 1A, 1B, W1, W2
        """
        if not self.calculation:
            return []

        fields = [
            LodgementField(
                field_code="G1",
                field_description="Total sales (including any GST)",
                amount=self._round_to_whole_dollars(self.calculation.g1_total_sales),
            ),
            LodgementField(
                field_code="G2",
                field_description="Export sales",
                amount=self._round_to_whole_dollars(self.calculation.g2_export_sales),
            ),
            LodgementField(
                field_code="G3",
                field_description="Other GST-free sales",
                amount=self._round_to_whole_dollars(self.calculation.g3_gst_free_sales),
            ),
            LodgementField(
                field_code="G10",
                field_description="Capital purchases (including any GST)",
                amount=self._round_to_whole_dollars(self.calculation.g10_capital_purchases),
            ),
            LodgementField(
                field_code="G11",
                field_description="Non-capital purchases (including any GST)",
                amount=self._round_to_whole_dollars(self.calculation.g11_non_capital_purchases),
            ),
            LodgementField(
                field_code="1A",
                field_description="GST on sales",
                amount=self._round_to_whole_dollars(self.calculation.field_1a_gst_on_sales),
            ),
            LodgementField(
                field_code="1B",
                field_description="GST on purchases",
                amount=self._round_to_whole_dollars(self.calculation.field_1b_gst_on_purchases),
            ),
        ]

        # Add PAYG fields if applicable
        if self.calculation.w1_total_wages and Decimal(str(self.calculation.w1_total_wages)) > 0:
            fields.extend(
                [
                    LodgementField(
                        field_code="W1",
                        field_description="Total salary, wages and other payments",
                        amount=self._round_to_whole_dollars(self.calculation.w1_total_wages),
                    ),
                    LodgementField(
                        field_code="W2",
                        field_description="Amount withheld from payments shown at W1",
                        amount=self._round_to_whole_dollars(self.calculation.w2_amount_withheld),
                    ),
                ]
            )

        return fields

    def generate_lodgement_pdf(self) -> bytes:
        """Generate PDF with ATO-compliant lodgement summary section.

        Returns:
            PDF file as bytes with lodgement summary
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title with APPROVED FOR LODGEMENT stamp
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=6,
            alignment=1,  # Center
        )
        story.append(Paragraph("Business Activity Statement", title_style))

        # Approved stamp
        stamp_style = ParagraphStyle(
            "Stamp",
            parent=styles["Normal"],
            fontSize=12,
            textColor=colors.HexColor("#059669"),
            alignment=1,
            spaceAfter=12,
        )
        story.append(Paragraph("✓ APPROVED FOR LODGEMENT", stamp_style))
        story.append(Spacer(1, 5 * mm))

        # Organization Info
        info_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontSize=11,
            spaceAfter=4,
        )
        story.append(Paragraph(f"<b>Client:</b> {self.organization_name}", info_style))
        if self.abn:
            story.append(Paragraph(f"<b>ABN:</b> {self.abn}", info_style))
        story.append(Paragraph(f"<b>Period:</b> {self.period.display_name}", info_style))
        story.append(
            Paragraph(
                f"<b>Period Dates:</b> {format_date(self.period.start_date)} - {format_date(self.period.end_date)}",
                info_style,
            )
        )
        story.append(Paragraph(f"<b>Due Date:</b> {format_date(self.period.due_date)}", info_style))
        if self.approved_by_name:
            story.append(Paragraph(f"<b>Approved by:</b> {self.approved_by_name}", info_style))
        if self.session.approved_at:
            story.append(
                Paragraph(
                    f"<b>Approved on:</b> {self.session.approved_at.strftime('%d %b %Y %H:%M')}",
                    info_style,
                )
            )

        # Lodgement details (if lodged)
        if self.session.lodged_at:
            story.append(Spacer(1, 5 * mm))
            lodgement_style = ParagraphStyle(
                "Lodgement",
                parent=info_style,
                textColor=colors.HexColor("#059669"),
            )
            story.append(Paragraph("<b>— LODGEMENT DETAILS —</b>", lodgement_style))
            story.append(
                Paragraph(
                    f"<b>Lodged on:</b> {self.session.lodged_at.strftime('%d %b %Y %H:%M')}",
                    info_style,
                )
            )
            if self.session.lodgement_method:
                method_label = {
                    "ATO_PORTAL": "ATO Business Portal",
                    "XERO": "Lodged via Xero",
                    "OTHER": self.session.lodgement_method_description or "Other Method",
                }.get(self.session.lodgement_method, self.session.lodgement_method)
                story.append(Paragraph(f"<b>Lodgement Method:</b> {method_label}", info_style))
            if self.session.ato_reference_number:
                story.append(
                    Paragraph(
                        f"<b>ATO Reference:</b> {self.session.ato_reference_number}", info_style
                    )
                )
            if self.lodged_by_name:
                story.append(Paragraph(f"<b>Lodged by:</b> {self.lodged_by_name}", info_style))

        story.append(Spacer(1, 10 * mm))

        if self.calculation:
            # Lodgement Summary Section
            story.append(Paragraph("<b>LODGEMENT SUMMARY - ATO BAS Fields</b>", styles["Heading2"]))
            story.append(Paragraph("(All amounts rounded to whole dollars)", styles["Normal"]))
            story.append(Spacer(1, 5 * mm))

            lodgement_fields = self._build_lodgement_summary_data()

            # Build lodgement summary table
            summary_data = [["Field", "Description", "Amount ($)"]]
            for field in lodgement_fields:
                summary_data.append(
                    [
                        field.field_code,
                        field.field_description,
                        f"${field.amount:,}",
                    ]
                )

            # Add separator and totals
            summary_data.append(["", "", ""])
            gst_payable = self._round_to_whole_dollars(self.calculation.gst_payable)
            summary_data.append(
                [
                    "",
                    "Net GST (1A minus 1B)" if gst_payable >= 0 else "GST Refund (1B minus 1A)",
                    f"${abs(gst_payable):,}",
                ]
            )

            if (
                self.calculation.w2_amount_withheld
                and Decimal(str(self.calculation.w2_amount_withheld)) > 0
            ):
                payg = self._round_to_whole_dollars(self.calculation.w2_amount_withheld)
                summary_data.append(["", "PAYG withholding (W2)", f"${payg:,}"])

            total = self._round_to_whole_dollars(self.calculation.total_payable)
            if total >= 0:
                summary_data.append(["", "TOTAL AMOUNT PAYABLE TO ATO", f"${total:,}"])
            else:
                summary_data.append(["", "TOTAL REFUND FROM ATO", f"${abs(total):,}"])

            summary_table = Table(summary_data, colWidths=[25 * mm, 90 * mm, 35 * mm])
            summary_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        # Highlight totals row
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fef3c7")),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ]
                )
            )
            story.append(summary_table)
            story.append(Spacer(1, 15 * mm))

            # Add the full working paper below
            story.append(Paragraph("<b>SUPPORTING CALCULATIONS</b>", styles["Heading2"]))
            story.append(Spacer(1, 5 * mm))

            # GST Section (same as base exporter but with whole dollars)
            gst_data = [
                ["Field", "Description", "Amount"],
                [
                    "G1",
                    "Total Sales (including GST)",
                    format_whole_dollars(self.calculation.g1_total_sales),
                ],
                ["G2", "Export Sales", format_whole_dollars(self.calculation.g2_export_sales)],
                [
                    "G3",
                    "Other GST-Free Sales",
                    format_whole_dollars(self.calculation.g3_gst_free_sales),
                ],
                [
                    "G10",
                    "Capital Purchases",
                    format_whole_dollars(self.calculation.g10_capital_purchases),
                ],
                [
                    "G11",
                    "Non-Capital Purchases",
                    format_whole_dollars(self.calculation.g11_non_capital_purchases),
                ],
                ["", "", ""],
                [
                    "1A",
                    "GST on Sales",
                    format_whole_dollars(self.calculation.field_1a_gst_on_sales),
                ],
                [
                    "1B",
                    "GST on Purchases",
                    format_whole_dollars(self.calculation.field_1b_gst_on_purchases),
                ],
                [
                    "",
                    "Net GST Payable/(Refundable)",
                    format_whole_dollars(self.calculation.gst_payable),
                ],
            ]

            gst_table = Table(gst_data, colWidths=[25 * mm, 90 * mm, 35 * mm])
            gst_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            story.append(gst_table)

        else:
            story.append(Paragraph("No calculation data available.", styles["Normal"]))

        # Footer
        story.append(Spacer(1, 20 * mm))
        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
            alignment=1,
        )
        story.append(
            Paragraph(
                f"Generated by Clairo on {datetime.now(UTC).strftime('%d %b %Y %H:%M UTC')}",
                footer_style,
            )
        )

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_lodgement_excel(self) -> bytes:
        """Generate Excel with dedicated Lodgement Summary sheet.

        Returns:
            Excel file as bytes with Lodgement Summary as first sheet
        """
        wb = Workbook()

        # Create Lodgement Summary sheet (first sheet)
        ws_summary = wb.active
        ws_summary.title = "Lodgement Summary"
        self._build_lodgement_summary_sheet(ws_summary)

        # Create Working Paper sheet (second sheet)
        ws_working = wb.create_sheet("Working Paper")
        self._build_working_paper_sheet(ws_working)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _build_lodgement_summary_sheet(self, ws) -> None:
        """Build the Lodgement Summary sheet with ATO-compliant formatting."""
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        approved_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        total_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 1

        # Title
        ws.merge_cells("A1:C1")
        ws["A1"] = "BAS LODGEMENT SUMMARY"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        row += 1

        # Approved stamp
        ws.merge_cells("A2:C2")
        ws["A2"] = "✓ APPROVED FOR LODGEMENT"
        ws["A2"].font = Font(bold=True, size=12, color="059669")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].fill = approved_fill
        row += 2

        # Client Info
        info_data = [
            ("Client:", self.organization_name),
        ]
        if self.abn:
            info_data.append(("ABN:", self.abn))
        info_data.extend(
            [
                ("Period:", self.period.display_name),
                (
                    "Period Dates:",
                    f"{format_date(self.period.start_date)} - {format_date(self.period.end_date)}",
                ),
                ("Due Date:", format_date(self.period.due_date)),
            ]
        )
        if self.approved_by_name:
            info_data.append(("Approved by:", self.approved_by_name))
        if self.session.approved_at:
            info_data.append(("Approved on:", self.session.approved_at.strftime("%d %b %Y %H:%M")))

        # Lodgement details (if lodged)
        if self.session.lodged_at:
            info_data.append(("", ""))  # Blank row
            info_data.append(("— LODGEMENT DETAILS —", ""))
            info_data.append(("Lodged on:", self.session.lodged_at.strftime("%d %b %Y %H:%M")))
            if self.session.lodgement_method:
                method_label = {
                    "ATO_PORTAL": "ATO Business Portal",
                    "XERO": "Lodged via Xero",
                    "OTHER": self.session.lodgement_method_description or "Other Method",
                }.get(self.session.lodgement_method, self.session.lodgement_method)
                info_data.append(("Lodgement Method:", method_label))
            if self.session.ato_reference_number:
                info_data.append(("ATO Reference:", self.session.ato_reference_number))
            if self.lodged_by_name:
                info_data.append(("Lodged by:", self.lodged_by_name))

        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        if self.calculation:
            # ATO Fields Header
            ws.merge_cells(f"A{row}:C{row}")
            ws.cell(row=row, column=1, value="ATO BAS FIELDS (Whole Dollars)").font = Font(
                bold=True, size=12
            )
            row += 1

            # Table Header
            for col, header in enumerate(["Field", "Description", "Amount ($)"], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            # Lodgement fields
            lodgement_fields = self._build_lodgement_summary_data()
            for field in lodgement_fields:
                ws.cell(row=row, column=1, value=field.field_code).border = thin_border
                ws.cell(row=row, column=2, value=field.field_description).border = thin_border
                cell = ws.cell(row=row, column=3, value=field.amount)
                cell.number_format = '"$"#,##0'
                cell.border = thin_border
                row += 1

            # Separator row
            row += 1

            # Totals
            gst_payable = self._round_to_whole_dollars(self.calculation.gst_payable)
            ws.cell(row=row, column=1, value="").border = thin_border
            ws.cell(
                row=row,
                column=2,
                value="Net GST (1A minus 1B)" if gst_payable >= 0 else "GST Refund (1B minus 1A)",
            ).border = thin_border
            ws.cell(row=row, column=2).font = Font(bold=True)
            cell = ws.cell(row=row, column=3, value=abs(gst_payable))
            cell.number_format = '"$"#,##0'
            cell.border = thin_border
            cell.font = Font(bold=True)
            row += 1

            if (
                self.calculation.w2_amount_withheld
                and Decimal(str(self.calculation.w2_amount_withheld)) > 0
            ):
                payg = self._round_to_whole_dollars(self.calculation.w2_amount_withheld)
                ws.cell(row=row, column=1, value="").border = thin_border
                ws.cell(row=row, column=2, value="PAYG withholding (W2)").border = thin_border
                ws.cell(row=row, column=2).font = Font(bold=True)
                cell = ws.cell(row=row, column=3, value=payg)
                cell.number_format = '"$"#,##0'
                cell.border = thin_border
                cell.font = Font(bold=True)
                row += 1

            # Total row
            total = self._round_to_whole_dollars(self.calculation.total_payable)
            ws.cell(row=row, column=1, value="").border = thin_border
            ws.cell(row=row, column=1).fill = total_fill
            label = "TOTAL AMOUNT PAYABLE TO ATO" if total >= 0 else "TOTAL REFUND FROM ATO"
            ws.cell(row=row, column=2, value=label).border = thin_border
            ws.cell(row=row, column=2).font = Font(bold=True, size=11)
            ws.cell(row=row, column=2).fill = total_fill
            cell = ws.cell(row=row, column=3, value=abs(total))
            cell.number_format = '"$"#,##0'
            cell.border = thin_border
            cell.font = Font(bold=True, size=11)
            cell.fill = total_fill

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 18

    def _build_working_paper_sheet(self, ws) -> None:
        """Build the Working Paper sheet (same as base exporter)."""
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 1

        # Title
        ws.merge_cells("A1:C1")
        ws["A1"] = "BAS Working Paper - Detailed Calculations"
        ws["A1"].font = Font(bold=True, size=14)
        row += 2

        if self.calculation:
            # GST Section
            ws.cell(row=row, column=1, value="GST Calculation").font = Font(bold=True, size=12)
            row += 1

            for col, header in enumerate(["Field", "Description", "Amount"], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            gst_data = [
                ("G1", "Total Sales (including GST)", self.calculation.g1_total_sales),
                ("G2", "Export Sales", self.calculation.g2_export_sales),
                ("G3", "Other GST-Free Sales", self.calculation.g3_gst_free_sales),
                ("G10", "Capital Purchases", self.calculation.g10_capital_purchases),
                ("G11", "Non-Capital Purchases", self.calculation.g11_non_capital_purchases),
                ("", "", None),
                ("1A", "GST on Sales", self.calculation.field_1a_gst_on_sales),
                ("1B", "GST on Purchases", self.calculation.field_1b_gst_on_purchases),
                ("", "Net GST Payable/(Refundable)", self.calculation.gst_payable),
            ]

            for field, desc, amount in gst_data:
                ws.cell(row=row, column=1, value=field).border = thin_border
                ws.cell(row=row, column=2, value=desc).border = thin_border
                cell = ws.cell(row=row, column=3, value=float(amount) if amount else 0)
                cell.number_format = '"$"#,##0.00'
                cell.border = thin_border
                row += 1

            row += 1

            # PAYG Section (if applicable)
            if (
                self.calculation.w1_total_wages
                and Decimal(str(self.calculation.w1_total_wages)) > 0
            ):
                ws.cell(row=row, column=1, value="PAYG Withholding").font = Font(bold=True, size=12)
                row += 1

                payg_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
                for col, header in enumerate(["Field", "Description", "Amount"], 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = payg_fill
                    cell.border = thin_border
                row += 1

                payg_data = [
                    (
                        "W1",
                        "Total Salary, Wages and Other Payments",
                        self.calculation.w1_total_wages,
                    ),
                    ("W2", "Amount Withheld from Payments", self.calculation.w2_amount_withheld),
                ]

                for field, desc, amount in payg_data:
                    ws.cell(row=row, column=1, value=field).border = thin_border
                    ws.cell(row=row, column=2, value=desc).border = thin_border
                    cell = ws.cell(row=row, column=3, value=float(amount) if amount else 0)
                    cell.number_format = '"$"#,##0.00'
                    cell.border = thin_border
                    row += 1

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 18

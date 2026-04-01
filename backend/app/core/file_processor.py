"""Shared file processing for chat attachments.

Handles storage (MinIO) and content extraction for files attached to
chat messages across all modules (AI Assistant, Tax Planning, Feedback).

Extracted content is sent to Claude alongside the user's text message;
the original file is stored in MinIO for persistence.

Supported formats:
- Images (PNG, JPEG, WebP, GIF) — sent as vision content blocks
- PDFs — sent as document content blocks
- CSV — parsed to text table
- Excel (.xlsx) — parsed with openpyxl to text table
"""

import base64
import csv
import io
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import UUID

from fastapi import UploadFile
from minio import Minio

from app.config import MinioSettings

logger = logging.getLogger(__name__)

# Allowed MIME types → category mapping
SUPPORTED_TYPES: dict[str, str] = {
    # Images — Claude vision
    "image/png": "image",
    "image/jpeg": "image",
    "image/webp": "image",
    "image/gif": "image",
    # PDF — Claude document
    "application/pdf": "pdf",
    # Spreadsheets — parsed to text
    "text/csv": "csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "excel",
    # Plain text
    "text/plain": "text",
}

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


@dataclass
class ProcessedFile:
    """Result of processing an uploaded file."""

    object_key: str  # MinIO storage key
    filename: str
    media_type: str
    category: str  # image, pdf, csv, excel, text
    size_bytes: int
    content_blocks: list[dict[str, Any]]  # Anthropic API content blocks


def _get_minio_client() -> tuple[Minio, str]:
    """Create a MinIO client from settings."""
    settings = MinioSettings()
    client = Minio(
        settings.endpoint,
        access_key=settings.access_key,
        secret_key=settings.secret_key.get_secret_value(),
        secure=settings.use_ssl,
    )
    return client, settings.bucket


def _parse_csv_to_text(data: bytes) -> str:
    """Parse CSV bytes into a markdown table string."""
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return "(empty CSV)"

    # Build markdown table
    header = rows[0]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in rows[1:101]:  # Cap at 100 data rows
        padded = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(padded[: len(header)]) + " |")

    if len(rows) > 102:
        lines.append(f"\n... ({len(rows) - 102} more rows truncated)")

    return "\n".join(lines)


def _parse_excel_to_text(data: bytes) -> str:
    """Parse Excel bytes into a markdown table string."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets_text = []

    for sheet_name in wb.sheetnames[:5]:  # Cap at 5 sheets
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(max_row=101, values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            continue

        header = rows[0]
        lines = [
            f"**Sheet: {sheet_name}**",
            "| " + " | ".join(header) + " |",
            "| " + " | ".join("---" for _ in header) + " |",
        ]
        for row in rows[1:]:
            lines.append("| " + " | ".join(row[: len(header)]) + " |")

        if ws.max_row and ws.max_row > 101:
            lines.append(f"\n... ({ws.max_row - 101} more rows truncated)")

        sheets_text.append("\n".join(lines))

    wb.close()
    return "\n\n".join(sheets_text) if sheets_text else "(empty spreadsheet)"


async def process_chat_attachment(
    file: UploadFile,
    tenant_id: UUID,
    module: str,
    context_id: UUID | str,
    message_id: str,
) -> ProcessedFile:
    """Process and store a chat file attachment.

    Stores the file in MinIO, then builds the appropriate Anthropic API
    content blocks based on file type.

    Args:
        file: The uploaded file.
        tenant_id: Tenant ID for storage path.
        module: Module name for storage path (e.g. 'tax-planning', 'assistant', 'feedback').
        context_id: Context-specific ID (plan_id, conversation_id, submission_id).
        message_id: Message identifier for storage path.

    Returns:
        ProcessedFile with storage key and content blocks.

    Raises:
        ValueError: If file type is unsupported or file is too large.
    """
    content_type = file.content_type or "application/octet-stream"
    category = SUPPORTED_TYPES.get(content_type)

    if not category:
        # Try to infer from extension
        ext = Path(file.filename or "").suffix.lower()
        ext_map = {
            ".png": ("image/png", "image"),
            ".jpg": ("image/jpeg", "image"),
            ".jpeg": ("image/jpeg", "image"),
            ".webp": ("image/webp", "image"),
            ".gif": ("image/gif", "image"),
            ".pdf": ("application/pdf", "pdf"),
            ".csv": ("text/csv", "csv"),
            ".xlsx": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "excel",
            ),
            ".txt": ("text/plain", "text"),
        }
        if ext in ext_map:
            content_type, category = ext_map[ext]
        else:
            supported = ", ".join(sorted({Path(f".{k}").suffix for k in ext_map}))
            raise ValueError(f"Unsupported file type: {content_type}. Supported: {supported}")

    # Read file content
    data = await file.read()
    size = len(data)

    if size > MAX_FILE_SIZE:
        raise ValueError(f"File too large ({size // 1024 // 1024}MB). Maximum is 10MB.")

    if size == 0:
        raise ValueError("File is empty.")

    # Store in MinIO (best-effort — skip if unavailable in prod)
    ext = Path(file.filename or "file").suffix.lower() or ".bin"
    object_key = f"{module}/{tenant_id}/{context_id}/chat/{message_id}{ext}"

    try:
        minio_client, bucket = _get_minio_client()
        minio_client.put_object(
            bucket,
            object_key,
            io.BytesIO(data),
            size,
            content_type=content_type,
        )
        logger.info(
            "Stored chat attachment: %s (%s, %d bytes)",
            object_key,
            content_type,
            size,
        )
    except Exception:
        logger.warning(
            "MinIO unavailable, skipping file storage for %s (%d bytes). "
            "File content will still be sent to Claude.",
            file.filename,
            size,
        )
        object_key = ""  # No storage key — file not persisted

    # Build Anthropic content blocks
    content_blocks: list[dict[str, Any]] = []

    if category == "image":
        b64 = base64.standard_b64encode(data).decode("ascii")
        content_blocks.append(
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": content_type,
                    "data": b64,
                },
            }
        )

    elif category == "pdf":
        b64 = base64.standard_b64encode(data).decode("ascii")
        content_blocks.append(
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": b64,
                },
            }
        )

    elif category == "csv":
        table_text = _parse_csv_to_text(data)
        content_blocks.append(
            {
                "type": "text",
                "text": f"[Attached CSV: {file.filename}]\n\n{table_text}",
            }
        )

    elif category == "excel":
        table_text = _parse_excel_to_text(data)
        content_blocks.append(
            {
                "type": "text",
                "text": f"[Attached Excel: {file.filename}]\n\n{table_text}",
            }
        )

    elif category == "text":
        text_content = data.decode("utf-8", errors="replace")[:50000]
        content_blocks.append(
            {
                "type": "text",
                "text": f"[Attached file: {file.filename}]\n\n{text_content}",
            }
        )

    return ProcessedFile(
        object_key=object_key,
        filename=file.filename or "file",
        media_type=content_type,
        category=category,
        size_bytes=size,
        content_blocks=content_blocks,
    )


def build_content_blocks_from_metadata(
    metadata: dict,
    minio_client: Minio | None = None,
    bucket: str | None = None,
) -> list[dict[str, Any]] | None:
    """Rebuild Anthropic content blocks from stored metadata.

    Used when replaying conversation history — re-reads the file from
    MinIO and constructs the appropriate content blocks.

    Returns None if no attachment in metadata.
    """
    attachment = metadata.get("attachment")
    if not attachment:
        return None

    object_key = attachment["object_key"]
    category = attachment["category"]
    media_type = attachment["media_type"]
    filename = attachment.get("filename", "file")

    if minio_client is None:
        minio_client, bucket = _get_minio_client()

    try:
        response = minio_client.get_object(bucket, object_key)
        data = response.read()
        response.close()
        response.release_conn()
    except Exception:
        logger.warning("Failed to read attachment from MinIO: %s", object_key)
        return [{"type": "text", "text": f"[Attachment: {filename} — file unavailable]"}]

    if category == "image":
        b64 = base64.standard_b64encode(data).decode("ascii")
        return [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": b64,
                },
            }
        ]

    if category == "pdf":
        b64 = base64.standard_b64encode(data).decode("ascii")
        return [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": b64,
                },
            }
        ]

    if category == "csv":
        table_text = _parse_csv_to_text(data)
        return [{"type": "text", "text": f"[Attached CSV: {filename}]\n\n{table_text}"}]

    if category == "excel":
        table_text = _parse_excel_to_text(data)
        return [{"type": "text", "text": f"[Attached Excel: {filename}]\n\n{table_text}"}]

    # text / fallback
    text_content = data.decode("utf-8", errors="replace")[:50000]
    return [{"type": "text", "text": f"[Attached file: {filename}]\n\n{text_content}"}]
